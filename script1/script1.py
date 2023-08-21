import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
def run_command(command):
    result = subprocess.run(command, stdout=subprocess.PIPE, text=True, shell=True)
    return result.stdout.splitlines()


commands = {
    "Netstat -ano": run_command("netstat -ano"),
    "Whoami": run_command("whoami"),
    "Systeminfo": run_command("systeminfo"),
    "ver": run_command("ver")
}

wb = Workbook()
ws = wb.active

bold_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
row_fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
row_fill_toggle = PatternFill(start_color="00808080", end_color="00808080", fill_type="solid")
center_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

ws.append(["Komut", "Sonuç", "", "", "Bilgi"])
for cell in ws["1:1"]:
    cell.font = bold_font
    cell.alignment = center_alignment
    cell.fill = header_fill

ws.merge_cells("B1:D1")

guidance_font = Font(color="000000")
guidance_fill = PatternFill(start_color="FFFFD9", end_color="FFFFD9", fill_type="solid")
guidance_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
guidance_row_height = 40
guidance_message = ("Bu program çeşitli komutların çıktılarını gösterir. "
                    "Tüm verileri görmek için sol taraftaki '+' butonunu kullanabilirsiniz.")
ws["E1"].value = guidance_message
ws["E1"].font = guidance_font
ws["E1"].fill = guidance_fill
ws["E1"].alignment = guidance_alignment
ws.row_dimensions[1].height = guidance_row_height

ws.merge_cells("E1:F1")

toggle = False
for command, result_lines in commands.items():
    start_row = ws.max_row + 1
    for idx, line in enumerate(result_lines, start=start_row):
        # Split lines based on multiple spaces
        parts = [part for part in line.split('  ') if part.strip() != '']
        if len(parts) > 1:
            ws.append([command if idx == start_row else ""] + parts)
        else:
            ws.append([command if idx == start_row else "", line])

    ws.merge_cells(start_row=start_row, end_row=start_row + len(result_lines) - 1, start_column=1, end_column=1)
    for cell in ws[f"A{start_row}:A{start_row + len(result_lines) - 1}"]:
        cell[0].alignment = center_alignment
        cell[0].fill = row_fill if toggle else row_fill_toggle

    toggle = not toggle

    if len(result_lines) > 10:
        ws.row_dimensions.group(start_row + 10, start_row + len(result_lines) - 1, outline_level=1, hidden=True)

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 20

wb.save("script1-çıktı.xlsx")